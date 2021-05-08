from pathlib import Path

from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from openpyxl import load_workbook

workbook = load_workbook(filename=Path(r"../test/Resources.xlsx"), read_only=True)

rooms_sheet = workbook['Rooms']
print(*rooms_sheet[1])
headers = [cell.value for cell in rooms_sheet[1]]
rows = [room for room in rooms_sheet.iter_rows(min_row=2, values_only=True)]


class TableModel(QAbstractTableModel):
    def rowCount(self, parent):
        return len(rows)

    def columnCount(self, parent):
        return len(headers)

    def data(self, index, role):
        if role != Qt.DisplayRole:
            return QVariant()
        return rows[index.row()][index.column()]

    def headerData(self, section, orientation, role):
        if role != Qt.DisplayRole or orientation != Qt.Horizontal:
            return QVariant()
        return headers[section]


if __name__ == '__main__':
    app = QApplication([])
    model = TableModel()
    view = QTableView()
    view.setModel(model)
    view.show()
    app.exec_()
