import time
from pathlib import Path
from pprint import pprint

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from Solver import *
from resources import *


def excel_export(promo_list: list[Promotion]):
    # saving solution to an excel file
    export_workbook = load_workbook(filename=Path(r"../test/result.xlsx"))

    for promo in promo_list:
        for sect in promo.list_section:
            EDT_sheet = export_workbook.copy_worksheet(export_workbook["template"])
            EDT_sheet.title = f'{promo.level}_{sect.number}'
            EDT_sheet["A5"] = f"Emploi du {promo.level}"
            nb_row = sect.nb_group
            for day in sect.EDT:
                for slot in day:
                    if len(slot.sessions) > nb_row:
                        nb_row = len(slot.sessions)
            # print(EDT_sheet[7][1].value)
            for j, day in enumerate(sect.EDT, start=1):
                for k, slot in enumerate(day, start=2):
                    slot.sessions.sort(key=lambda s: s.attendance.number)
                    for session_index, session in enumerate(slot.sessions, start=1):
                        # print(j * 7 + i,k)
                        nb_row = max(9, sect.nb_group)
                        start_session_index = (j - 1) * nb_row + session_index + 7
                        end_session_index = j * nb_row + session_index - 1 + 7
                        if nb_row > 9:
                            EDT_sheet.insert_rows(end_session_index - 2, nb_row - 9)
                        if session.session_type == SessionType.Cour:
                            EDT_sheet.merge_cells(start_row=start_session_index, start_column=k,
                                                  end_row=end_session_index,
                                                  end_column=k)
                            EDT_sheet.cell(start_session_index, k).font = Font(bold=True)
                            EDT_sheet.cell(start_session_index, k).alignment = Alignment(wrap_text=True,
                                                                                         horizontal='center',
                                                                                         vertical='center')
                        EDT_sheet.cell(start_session_index, k).value = str(session)
    del export_workbook["template"]
    export_workbook.save('FSEI_Mosta_EDT.xlsx')


def get_data(xlsx):
    prof_names = ['Abbassa',
                  'Abbes',
                  'ABID M.',
                  'Adda',
                  'BAHNES N.',
                  'Beghdad',
                  'Belayachi',
                  'Belhachemi',
                  'Belhakem',
                  'belhalfaoui',
                  'Belhouari',
                  'Belouatek',
                  'BENAMEUR',
                  'Bencherif',
                  'Benguettat',
                  'BENHAMED',
                  'BENIDRIS F.Z',
                  'Benmalti',
                  'Benotsmane',
                  'BENSALLOUA',
                  'BENTAOUZA C',
                  'BESNASSI',
                  'BESSAOUD K.',
                  'BETOUATI',
                  'Bouattou',
                  'Boukra',
                  'Boulenouar',
                  'Bourada',
                  'Bourahla',
                  'BOUZEBIBA',
                  'DELALI A.',
                  'DJAHAFI',
                  'DJEBBARA R.',
                  'FILALI F.',
                  'HABIB ZAHMANI M',
                  'HAMAMI',
                  'Hamiani',
                  'harrats',
                  'HASSAINE',
                  'HENNI F',
                  'HENNI K',
                  'HENNI K.',
                  'Hentit H',
                  'Hentit N',
                  'HOCINE N.',
                  'Kadi',
                  'KAID SLIMANE',
                  'KENNICHE A.',
                  'KHELIFA N.',
                  'KHIAT',
                  'Labdelli',
                  'LAREDJ A. M',
                  'Latreuch',
                  'M. AMIR A.',
                  'M. Andasmas M',
                  'M. B. BENDOUKHA',
                  'M. BELAIDI Benharrat',
                  'M. BELARBI Lakehal',
                  'M. Belhamiti Omar',
                  'M. Benchehida',
                  'M. Benyatou Kamel',
                  'M. Benzidane',
                  'M. BOUZIT Hamid',
                  'M. Dahmani Z.',
                  'M. Fettouch Houari',
                  'M. Ghezzar Med',
                  'M. Kaid',
                  'M. Kaid Med',
                  'M. Medeghri Ahmed',
                  'M. Menad Abdallah',
                  'M. Mohammedi Mustapha',
                  'M. Ould Ali M',
                  'M. Zoubir DAHMANI',
                  'M.Ould ali',
                  'MAGHNI SANDID Z.',
                  'MECHAOUI M. D.G',
                  'Meghoufel',
                  'Mekemmeche',
                  'Melati',
                  'MENAD',
                  'MEROUFEL B.',
                  'Messaoudi',
                  'MIDOUN M.',
                  'MIROUD',
                  'Mlle Ali Merina.H',
                  'Mlle Amina FERRAOUN',
                  'Mlle Benaouad',
                  'Mlle Bouzid',
                  'Mlle Bouzid Leila',
                  'Mlle Dj. BENSIKADDOUR',
                  'Mlle Dj.BENSIKADDOUR',
                  'Mlle FERRAOUN A.',
                  'Mlle Hamou Maamar.M',
                  'Mlle Lakeb Ouda',
                  'Mlle Medeghri Ahmed',
                  'Mlle Zelmat Souhila',
                  'Mme Ablaoui',
                  'Mme Belmouhoub-Ould ali',
                  'Mme Bendahmane Hafida',
                  'Mme Bendehmane H',
                  'Mme Bouabdelli',
                  'Mme Diala.H',
                  'Mme Kaisserli',
                  'Mme Limam',
                  'Mme SAIDANI',
                  'Mme TABHARIT',
                  'Mme.Bechaoui',
                  'MOUMEN M.',
                  'MOUSSA M.',
                  'Mr Ablaoui H.',
                  'Mr Bouzit H',
                  'Mr S. M. Bahri',
                  'Mr. BOUAGADA',
                  'Mr.Senouci',
                  'SEHABA K.',
                  'Terki']
    modules_dict = {}
    promos_dict = {}
    professors_dict = {}
    # using the data in the test file to check the validity of the data model
    # the code looks bad but it's only test code
    promo_data = []
    prof_data = []
    roomu_data = []
    modules_data = []
    module_assignment_data = []
    workbook = load_workbook(filename=Path(xlsx), read_only=True)
    Module_sheet = workbook['Modules']
    promos_sheet = workbook['Promos']
    rooms_sheet = workbook['Rooms']
    for promo_row in promos_sheet.iter_rows(min_row=2, values_only=True):
        promo_name, nb_section, nb_group, group_effective, mod_start, mod_end = promo_row
        promo_data.append({"Name": promo_name, "Number of Sections": int(nb_section), "Number of Groups": int(nb_group),
                           "Effective per Group": int(group_effective)})

        promos_dict[promo_name] = Promotion(promo_name)
        promos_dict[promo_name].list_section = [Section(i + 1) for i in range(nb_section)]
        for section in promos_dict[promo_name].list_section:
            section.list_group = [Group(i + 1, group_effective) for i in range(nb_group // nb_section)]
            nb_group -= nb_group // nb_section
            nb_section -= 1
        modules = []
        modules_ass = []
        for module_row in Module_sheet.iter_rows(min_row=mod_start, max_row=mod_end, min_col=2, values_only=True):
            specific_mod = []
            # print(module_row)
            modules.append({"Name": module_row[0], "abriv": module_row[1], "nb_cour": module_row[2] or 0,
                            "nb_td": module_row[3] or 0,
                            "nb_tp": module_row[4] or 0})
            mod = Module(module_row[0], module_row[1], module_row[2] or 0, module_row[3] or 0, module_row[4] or 0)
            modules_dict[module_row[0]] = mod
            cour_str, td_str, tp_str = module_row[5], module_row[6], module_row[7]
            if cour_str:
                cour_profs_strings = cour_str.split(",")
                cour_profs = []
                for prof_string in cour_profs_strings:
                    name = prof_string.split("(")[0]

                    times = int(prof_string[prof_string.find("(") + 1].split(")")[0])
                    prof_data.append(name)
                    specific_mod.append({"prof_name": prof_names.index(name), "number": times, "type": 1})
                    cour_profs.extend([Professor(name) for _ in range(times * mod.nb_cour)])
                for sect_i in promos_dict[promo_name].list_section:
                    sect_i.add_required_sessions(
                        [(cour_profs.pop(), sect_i, mod, SessionType.Cour) for _ in range(mod.nb_cour)])
            if td_str:
                td_profs_strings = td_str.split(",")
                td_profs = []
                for prof_string in td_profs_strings:
                    name = prof_string.split("(")[0]
                    times = int(prof_string[prof_string.find("(") + 1].split(")")[0])
                    specific_mod.append({"prof_name": prof_names.index(name.strip()), "number": times, "type": 2})
                    prof_data.append(name)
                    td_profs.extend([Professor(name) for _ in range(times * mod.nb_td)])
                for sect_i in promos_dict[promo_name].list_section:
                    for group in sect_i.list_group:
                        sect_i.add_required_sessions(
                            [(td_profs.pop(), group, mod, SessionType.Td) for _ in range(mod.nb_td)])
            if tp_str:
                tp_profs_strings = tp_str.split(",")
                tp_profs = []
                for prof_string in tp_profs_strings:
                    name = prof_string.split("(")[0]
                    prof_data.append(name)
                    times = int(prof_string[prof_string.find("(") + 1].split(")")[0])
                    specific_mod.append({"prof_name": prof_names.index(name.strip()), "number": times, "type": 3})
                    tp_profs.extend([Professor(name) for _ in range(times * mod.nb_tp)])
                for sect_i in promos_dict[promo_name].list_section:
                    for group in sect_i.list_group:
                        sect_i.add_required_sessions(
                            [(tp_profs.pop(), group, mod, SessionType.Tp) for _ in range(mod.nb_tp)])
                    sect_i.required_sessions.sort(key=lambda s: s[3].value)
            modules_ass.append(specific_mod)
        module_assignment_data.append(modules_ass)
        modules_data.append(modules)
    # *room == room_name, room_type, capacity = room
    rooms_list = [(Room(*room)) for room in rooms_sheet.iter_rows(min_row=2, values_only=True)]
    for room in rooms_sheet.iter_rows(min_row=2, values_only=True):
        roomu_data.append({"Name": room[0], "Capacity": int(room[2]), "RoomType": int(room[1])})

    prof_data = [s.strip() for s in prof_data]
    prof_data = list(set(prof_data))
    prof_data.sort(key=lambda x: x.lower())
    #pprint(module_assignment_data)
    return modules_dict, promos_dict, professors_dict, rooms_list


if __name__ == '__main__':
    start = time.perf_counter()
    modules, promos, professors, rooms = get_data(r"../test/Resources.xlsx")
    promos = list(promos.values())
    data_shows = [DataShow(promos[:10]) for _ in range(4)]
    print([p.level for p in promos[:10]])
    data_shows.extend([DataShow(promos[13:15]) for _ in range(2)])
    print([p.level for p in promos[13:15]])
    physic = promos[10:13] + [promos[-1]]
    print([p.level for p in physic])
    data_shows.extend([DataShow(physic) for _ in range(2)])
    fsei_mosta = Faculty("FSEI-MOSTA", promos, rooms, data_shows)

    end = time.perf_counter()
    print("Time for import from excel +:", format((end - start) * 1000, ".2f"), "ms")
    start = time.perf_counter()

    # promos.reverse()
    problem_emploi_du_temp = PET(fsei_mosta)
    number_assignments = sum(len(session_list) for session_list in problem_emploi_du_temp.sessions_list)
    problem_emploi_du_temp.add_hard_constraint(ProfessorAvailability())
    problem_emploi_du_temp.add_hard_constraint(StudentAvailability())
    problem_emploi_du_temp.add_hard_constraint(RoomAvailability())
    problem_emploi_du_temp.add_hard_constraint(ThreeConsecutiveMaxSessions())
    problem_emploi_du_temp.add_hard_constraint(TwoCourPerDayMax())
    # very inefficient to add this
    problem_emploi_du_temp.add_hard_constraint(UniqueSessionDaily())
    #

    if problem_emploi_du_temp.solve():
        end = time.perf_counter()
        print("successfully generated EDT(", number_assignments, " assignments) in",
              format((end - start) * 1000, ".2f"), "ms")
    else:
        print("nope debug more")
    # promos.reverse()
    start = time.perf_counter()
    # excel_export(list(promos))
    end = time.perf_counter()
    print("exported to excel in", format((end - start) * 1000, ".2f"), "ms")
